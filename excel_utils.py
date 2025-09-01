import os
import pandas as pd
from openpyxl import load_workbook
from helpers import ek_ekle, format_number

KATSAyi_DOSYASI = "data/katsayılar.xlsx"
EMEKLI_SABLON = "data/emekli.xlsx"

def emekli_excel_olustur(ad_soyad, dairesi, ay, yil, tcno, gorev, save_path, iban, mudurluk, derece, kademe,
                         em_tarih, duzenleyen_ad, duzenleyen_unvan,
                         ger_gor_ad, ger_gor_unvan, har_yet_ad, har_yet_unvan):
    """Emekli yolluğu için Excel dosyası oluşturur ve kaydeder."""
    if not (tcno.isdigit() and len(tcno) == 11):
        raise ValueError("TC Kimlik Numarası 11 haneli olmalıdır!")

    if not os.path.exists(KATSAyi_DOSYASI):
        raise FileNotFoundError("katsayılar.xlsx bulunamadı!")

    df = pd.read_excel(KATSAyi_DOSYASI, index_col=0)
    if yil not in df.index or ay not in df.columns:
        raise ValueError(f"{yil} yılı {ay}. ay için katsayı bulunamadı!")

    katsayi = df.loc[yil, ay]
    gosterge = 13558
    damga = 0.00759
    der_kad = f"{derece}/{kademe}"

    toplam = round(float(katsayi) * gosterge, 2)
    vergi = round(toplam * damga, 2)
    odenecek = round(toplam - vergi, 2)
    ek = ek_ekle(ad_soyad)

    if not os.path.exists(EMEKLI_SABLON):
        raise FileNotFoundError("emekli.xlsx bulunamadı!")

    wb = load_workbook(EMEKLI_SABLON)
    ws = wb.active

    # Verileri ilgili hücrelere yaz
    ws["B4"] = mudurluk
    ws["J6"] = dairesi
    ws["J7"] = yil
    ws["C6"] = katsayi
    ws["C7"] = gosterge
    ws["B9"] = tcno
    ws["C9"] = ad_soyad
    ws["D9"] = gorev
    ws["E9"] = em_tarih
    ws["F9"] = der_kad
    ws["G9"] = iban
    ws["H9"] = toplam
    ws["I9"] = vergi
    ws["J9"] = odenecek

    ws["B21"] = duzenleyen_ad
    ws["B22"] = duzenleyen_unvan
    ws["D21"] = ger_gor_ad
    ws["D22"] = ger_gor_unvan
    ws["I21"] = har_yet_ad
    ws["I22"] = har_yet_unvan

    ws["B11"] = (f"{ad_soyad}'{ek} emekli olmasından dolayı "
                f"{format_number(toplam, 2)} Türk lirası tazminatın damga vergisi düşülerek "
                f"{format_number(odenecek, 2)} Türk Lirası ödenecektir.")

    

    wb.save(save_path)
    return save_path
